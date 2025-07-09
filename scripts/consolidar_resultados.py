import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

# Carpeta de resultados
RESULTS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../results'))
OUTPUT_FILE = os.path.abspath(os.path.join(RESULTS_DIR, 'consolidado_resultados.xlsx'))

# Buscar archivos .jtl y .csv
jtl_files = glob.glob(os.path.join(RESULTS_DIR, '*.jtl'))
csv_files = glob.glob(os.path.join(RESULTS_DIR, '*-agg.csv'))

# Función para extraer métricas clave de un archivo JMeter .jtl
# (asume formato XML o CSV, aquí se asume CSV para simplificar)
def extraer_metricas_jtl(filepath):
    try:
        df = pd.read_csv(filepath)
    except Exception:
        # Si no es CSV, intentar como XML (no implementado aquí)
        return None
    if 'elapsed' not in df.columns:
        return None
    metrica = {
        'Archivo': os.path.basename(filepath),
        'Samples': len(df),
        'Avg Response Time (ms)': df['elapsed'].mean(),
        'Median (ms)': df['elapsed'].median(),
        '90th pct (ms)': df['elapsed'].quantile(0.9),
        '95th pct (ms)': df['elapsed'].quantile(0.95),
        '99th pct (ms)': df['elapsed'].quantile(0.99),
        'Min (ms)': df['elapsed'].min(),
        'Max (ms)': df['elapsed'].max(),
        'Throughput (req/s)': round(len(df) / (df['timeStamp'].max() - df['timeStamp'].min()) * 1000, 2) if 'timeStamp' in df.columns else None,
        'Error %': round(100 * (df['success'] == False).sum() / len(df), 2) if 'success' in df.columns else None,
    }
    return metrica

# Consolidar métricas de todos los archivos .jtl
metricas = []
for f in jtl_files:
    m = extraer_metricas_jtl(f)
    if m:
        metricas.append(m)

# Consolidar Aggregate Reports si existen
agg_dfs = []
for f in csv_files:
    try:
        agg = pd.read_csv(f)
        agg['Archivo'] = os.path.basename(f)
        agg_dfs.append(agg)
    except Exception:
        pass

# Crear Excel interpretativo
wb = Workbook()
ws = wb.active
ws.title = 'Resumen'

# Escribir resumen de métricas
if metricas:
    ws.append(list(metricas[0].keys()))
    for m in metricas:
        ws.append(list(m.values()))
    for cell in ws[1]:
        cell.font = Font(bold=True)
else:
    ws.append(['No se encontraron archivos .jtl válidos'])

# Agregar hojas para Aggregate Reports
for agg in agg_dfs:
    name = agg['Archivo'].iloc[0].replace('.csv','')[:31]
    ws_agg = wb.create_sheet(title=name)
    ws_agg.append(list(agg.columns))
    for row in agg.itertuples(index=False):
        ws_agg.append(list(row))
    for cell in ws_agg[1]:
        cell.font = Font(bold=True)

# Hoja de análisis interpretativo
ws_analysis = wb.create_sheet(title='Análisis')
ws_analysis.append(['Fecha', 'Escenario', 'Servicio', 'Hallazgos', 'Recomendaciones'])
ws_analysis.append([datetime.now().strftime('%Y-%m-%d'), '', '', 'Completa aquí tu análisis según la rúbrica', ''])
for cell in ws_analysis[1]:
    cell.font = Font(bold=True)

wb.save(OUTPUT_FILE)
print(f'Consolidación completada. Archivo generado: {OUTPUT_FILE}')
